"""Loader: NEVO 2025 xlsx -> Postgres.
Loader: NEVO 2025 xlsx into Postgres.

Usage:
    .venv/bin/python -m src.load_nevo NEVO2025_v9.0.xlsx
    .venv/bin/python -m src.load_nevo NEVO2025_v9.0.xlsx --no-truncate
    .venv/bin/python -m src.load_nevo NEVO2025_v9.0.xlsx --dry-run

Idempotent: TRUNCATE + reload in a single transaction.
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path
from typing import Iterator

import psycopg
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

load_dotenv()

SHEET_NUTRIENTS = "NEVO2025_Nutrienten_Nutrients"
SHEET_FOODS = "NEVO2025"
SHEET_DETAILS = "NEVO2025_Details"

NUTRIENT_COLS = {
    "group_nl": "Voedingsstofgroep",
    "group_en": "Component group",
    "code": "Nutrient-code",
    "name_nl": "Voedingsstof",
    "name_en": "Component",
    "unit": "Eenheid/Unit",
}

FOOD_COLS = {
    "food_group_nl": "Voedingsmiddelgroep",
    "food_group_en": "Food group",
    "nevo_code": "NEVO-code",
    "name_nl": "Voedingsmiddelnaam/Dutch food name",
    "name_en": "Engelse naam/Food name",
    "synonyms": "Synoniem",
    "quantity": "Hoeveelheid/Quantity",
    "note": "Opmerking",
    "contains_traces_of": "Bevat sporen van/Contains traces of",
    "is_fortified_with": "Is verrijkt met/Is fortified with",
}

DETAIL_COLS = {
    "nevo_code": "NEVO-code",
    "nutrient_code": "Nutrient-code",
    "value": "Gehalte/Value",
}


def parse_value(raw):
    """Cell -> numeric or None. Placeholders ('-', 'Sp.', 'tr', '') become None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _opt(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _req(value) -> str:
    return "" if value is None else str(value).strip()


def header_index(ws: Worksheet, expected: dict[str, str]) -> dict[str, int]:
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    found: dict[str, int] = {}
    missing = []
    for key, header_name in expected.items():
        try:
            found[key] = headers.index(header_name)
        except ValueError:
            missing.append(header_name)
    if missing:
        raise SystemExit(
            f"sheet {ws.title!r}: missing headers {missing!r}. got: {headers!r}"
        )
    return found


def parse_nutrients(ws: Worksheet) -> list[tuple]:
    """NEVO has a handful of nutrient codes (PROT, FAT, CHO, FIBT, ASH) listed
    twice — once under 'Energie en macronutriënten' and once under their own
    group. Our PK is `code`; we keep the first occurrence and log the dupes."""
    cols = header_index(ws, NUTRIENT_COLS)
    rows: list[tuple] = []
    seen: set[str] = set()
    dupes: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[cols["code"]] is None:
            continue
        code = _req(row[cols["code"]])
        if code in seen:
            dupes.append(code)
            continue
        seen.add(code)
        rows.append(
            (
                code,
                _req(row[cols["group_nl"]]),
                _req(row[cols["group_en"]]),
                _req(row[cols["name_nl"]]),
                _req(row[cols["name_en"]]),
                _req(row[cols["unit"]]),
            )
        )
    if dupes:
        print(f"[loader] dedup nutrients: dropped duplicate codes {dupes}")
    return rows


def parse_foods(ws: Worksheet) -> list[tuple]:
    cols = header_index(ws, FOOD_COLS)
    rows: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nevo = row[cols["nevo_code"]]
        if nevo is None:
            continue
        rows.append(
            (
                int(nevo),
                _req(row[cols["name_nl"]]),
                _req(row[cols["name_en"]]),
                _req(row[cols["food_group_nl"]]),
                _req(row[cols["food_group_en"]]),
                _opt(row[cols["synonyms"]]),
                _req(row[cols["quantity"]]),
                _opt(row[cols["note"]]),
                _opt(row[cols["contains_traces_of"]]),
                _opt(row[cols["is_fortified_with"]]),
            )
        )
    return rows


def iter_details(
    ws: Worksheet,
    valid_codes: set[int],
    valid_nutrients: set[str],
    stats: dict[str, int],
) -> Iterator[tuple]:
    """Stream (nevo_code, nutrient_code, value) tuples. Skips rows with unknown
    FK and de-duplicates by (nevo_code, nutrient_code) — NEVO repeats some
    rows under different nutrient groups (PROT/FAT/CHO/FIBT/ASH)."""
    cols = header_index(ws, DETAIL_COLS)
    seen: set[tuple[int, str]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        nevo = row[cols["nevo_code"]]
        nut = row[cols["nutrient_code"]]
        if nevo is None or nut is None:
            continue
        nevo_int = int(nevo)
        nut_str = str(nut).strip()
        if nevo_int not in valid_codes:
            stats["skipped_food"] += 1
            continue
        if nut_str not in valid_nutrients:
            stats["skipped_nut"] += 1
            continue
        key = (nevo_int, nut_str)
        if key in seen:
            stats["skipped_dup"] += 1
            continue
        seen.add(key)
        v = parse_value(row[cols["value"]])
        if v is None:
            stats["null"] += 1
        stats["written"] += 1
        yield (nevo_int, nut_str, v)


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Load NEVO xlsx into Postgres.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument(
        "--truncate",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="TRUNCATE tables before insert (default: yes)",
    )
    parser.add_argument("--dry-run", action="store_true", help="parse only, no DB writes")
    parser.add_argument("--db-url", default=os.environ.get("NEVO_LOADER_URL"))
    args = parser.parse_args(argv)

    if not args.xlsx.exists():
        raise SystemExit(f"file not found: {args.xlsx}")
    if not args.dry_run and not args.db_url:
        raise SystemExit("NEVO_LOADER_URL not set (and --db-url not given)")

    t0 = time.perf_counter()
    size_mb = args.xlsx.stat().st_size / 1e6
    print(f"[loader] reading xlsx: {args.xlsx} ({size_mb:.1f} MB)")
    wb = load_workbook(args.xlsx, read_only=True, data_only=True)

    nutrients = parse_nutrients(wb[SHEET_NUTRIENTS])
    print(f"[loader] parsed nutrients: {len(nutrients)} rows")

    foods = parse_foods(wb[SHEET_FOODS])
    print(f"[loader] parsed foods:     {len(foods)} rows")

    if args.dry_run:
        details_count = sum(
            1 for _ in wb[SHEET_DETAILS].iter_rows(min_row=2, values_only=True)
        )
        print(f"[loader] details rows (raw): {details_count}")
        print("[loader] dry-run: no DB writes")
        return 0

    valid_codes = {f[0] for f in foods}
    valid_nutrients = {n[0] for n in nutrients}
    stats = {
        "skipped_food": 0,
        "skipped_nut": 0,
        "skipped_dup": 0,
        "null": 0,
        "written": 0,
    }

    with psycopg.connect(args.db_url, autocommit=False) as conn, conn.cursor() as cur:
        if args.truncate:
            print("[loader] truncating tables...")
            cur.execute(
                "TRUNCATE food_nutrients, foods, nutrients RESTART IDENTITY CASCADE"
            )
        print(f"[loader] inserting nutrients: {len(nutrients)}")
        cur.executemany(
            "INSERT INTO nutrients (code, group_nl, group_en, name_nl, name_en, unit) "
            "VALUES (%s, %s, %s, %s, %s, %s)",
            nutrients,
        )
        print(f"[loader] inserting foods:     {len(foods)}")
        cur.executemany(
            "INSERT INTO foods ("
            "nevo_code, name_nl, name_en, food_group_nl, food_group_en, "
            "synonyms, quantity, note, contains_traces_of, is_fortified_with"
            ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            foods,
        )
        print("[loader] copying food_nutrients (streaming)...")
        copy_sql = (
            "COPY food_nutrients (nevo_code, nutrient_code, value_per_100) FROM STDIN"
        )
        with cur.copy(copy_sql) as cp:
            for row in iter_details(
                wb[SHEET_DETAILS], valid_codes, valid_nutrients, stats
            ):
                cp.write_row(row)
        conn.commit()

    dt = time.perf_counter() - t0
    print(
        f"[loader] food_nutrients written: {stats['written']} "
        f"({stats['null']} NULL, "
        f"{stats['skipped_dup']} skipped duplicate, "
        f"{stats['skipped_food']} skipped unknown food, "
        f"{stats['skipped_nut']} skipped unknown nutrient)"
    )
    print(f"[loader] commit OK in {dt:.1f}s")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
